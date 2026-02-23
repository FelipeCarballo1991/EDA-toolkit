"""
Unit tests for FileExporter class, focusing on export_tables() method.
Tests the unified interface for exporting single or multiple DataFrames.
"""

import pytest
import pandas as pd
from pathlib import Path
import openpyxl
from pandas_toolkit.io.exporter import FileExporter


class TestFileExporter:
    """Test suite for FileExporter"""
    
    @pytest.fixture
    def exporter(self, tmp_path):
        """Create a FileExporter with temporary output directory"""
        return FileExporter(output_dir=str(tmp_path), verbose=False)
    
    @pytest.fixture
    def sample_df(self):
        """Create a sample DataFrame"""
        return pd.DataFrame({
            'id': [1, 2, 3, 4, 5],
            'name': ['Alice', 'Bob', 'Charlie', 'David', 'Eve'],
            'value': [100, 200, 300, 400, 500]
        })
    
    @pytest.fixture
    def large_df(self):
        """Create a large DataFrame for testing splits (2000 rows)"""
        return pd.DataFrame({
            'id': range(2000),
            'value': range(2000),
            'category': ['A', 'B', 'C', 'D'] * 500
        })
    
    @pytest.fixture
    def multiple_dfs(self):
        """Create multiple DataFrames"""
        df1 = pd.DataFrame({'col1': [1, 2, 3], 'col2': ['a', 'b', 'c']})
        df2 = pd.DataFrame({'colA': [10, 20], 'colB': ['x', 'y']})
        df3 = pd.DataFrame({'data': [100, 200, 300, 400]})
        return [df1, df2, df3]


class TestExportTablesSingleDataFrame(TestFileExporter):
    """Tests for export_tables with single DataFrame"""
    
    def test_export_single_dataframe_to_excel(self, exporter, sample_df, tmp_path):
        """Test exporting a single DataFrame to Excel"""
        filename = "single.xlsx"
        exporter.export_tables(sample_df, filename=filename)
        
        # Verify file was created
        filepath = tmp_path / filename
        assert filepath.exists()
        
        # Verify content
        df_read = pd.read_excel(filepath, sheet_name="Sheet1")
        pd.testing.assert_frame_equal(df_read, sample_df)
    
    def test_export_single_dataframe_in_list(self, exporter, sample_df, tmp_path):
        """Test exporting [df] (list with single DataFrame)"""
        filename = "single_list.xlsx"
        exporter.export_tables([sample_df], filename=filename)
        
        filepath = tmp_path / filename
        assert filepath.exists()
        
        # Should create Sheet1
        df_read = pd.read_excel(filepath, sheet_name="Sheet1")
        pd.testing.assert_frame_equal(df_read, sample_df)
    
    def test_export_tables_auto_adds_extension(self, exporter, sample_df, tmp_path):
        """Test that .xlsx extension is added automatically"""
        filename = "no_extension"
        exporter.export_tables(sample_df, filename=filename)
        
        filepath = tmp_path / "no_extension.xlsx"
        assert filepath.exists()
    
    def test_export_tables_to_csv(self, exporter, sample_df, tmp_path):
        """Test exporting single DataFrame to CSV"""
        filename = "output.csv"
        exporter.export_tables(sample_df, filename=filename, method="csv")
        
        filepath = tmp_path / filename
        assert filepath.exists()
        
        df_read = pd.read_csv(filepath)
        pd.testing.assert_frame_equal(df_read, sample_df)


class TestExportTablesMultipleDataFrames(TestFileExporter):
    """Tests for export_tables with multiple DataFrames"""
    
    def test_export_multiple_tables_to_excel(self, exporter, multiple_dfs, tmp_path):
        """Test exporting multiple DataFrames to single Excel file"""
        filename = "multiple.xlsx"
        exporter.export_tables(multiple_dfs, filename=filename)
        
        filepath = tmp_path / filename
        assert filepath.exists()
        
        # Verify all sheets were created
        wb = openpyxl.load_workbook(filepath)
        assert len(wb.sheetnames) == 3
        assert "Table1" in wb.sheetnames
        assert "Table2" in wb.sheetnames
        assert "Table3" in wb.sheetnames
        
        # Verify content of each sheet
        df1_read = pd.read_excel(filepath, sheet_name="Table1")
        df2_read = pd.read_excel(filepath, sheet_name="Table2")
        df3_read = pd.read_excel(filepath, sheet_name="Table3")
        
        pd.testing.assert_frame_equal(df1_read, multiple_dfs[0])
        pd.testing.assert_frame_equal(df2_read, multiple_dfs[1])
        pd.testing.assert_frame_equal(df3_read, multiple_dfs[2])
    
    def test_export_multiple_tables_csv_raises_error(self, exporter, multiple_dfs):
        """Test that CSV export with multiple tables raises ValueError"""
        with pytest.raises(ValueError, match="CSV export only supports single table"):
            exporter.export_tables(multiple_dfs, filename="output.csv", method="csv")
    
    def test_export_empty_list_raises_error(self, exporter):
        """Test that exporting empty list raises ValueError"""
        with pytest.raises(ValueError, match="No tables provided"):
            exporter.export_tables([], filename="output.xlsx")


class TestExportTablesLargeDataFrames(TestFileExporter):
    """Tests for export_tables with large DataFrames that need splitting"""
    
    def test_export_large_dataframe_splits_into_sheets(self, exporter, tmp_path):
        """Test that large DataFrame is split into multiple sheets"""
        # Create DataFrame with 1500 rows (will split with max_rows=1000)
        large_df = pd.DataFrame({
            'id': range(1500),
            'value': range(1500)
        })
        
        filename = "large.xlsx"
        exporter.export_tables(
            large_df,
            filename=filename,
            max_rows_per_sheet=1000
        )
        
        filepath = tmp_path / filename
        assert filepath.exists()
        
        # Should create 2 sheets: Sheet1 (1000 rows), Sheet2 (500 rows)
        wb = openpyxl.load_workbook(filepath)
        assert len(wb.sheetnames) == 2
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames
        
        # Verify row counts
        df1 = pd.read_excel(filepath, sheet_name="Sheet1")
        df2 = pd.read_excel(filepath, sheet_name="Sheet2")
        
        assert len(df1) == 1000
        assert len(df2) == 500
        
        # Verify data integrity
        df_combined = pd.concat([df1, df2], ignore_index=True)
        pd.testing.assert_frame_equal(df_combined, large_df)
    
    def test_export_multiple_large_tables_with_splits(self, exporter, tmp_path):
        """Test exporting multiple tables where one needs splitting"""
        small_df = pd.DataFrame({'col': [1, 2, 3]})
        large_df = pd.DataFrame({'data': range(1500)})
        
        filename = "mixed.xlsx"
        exporter.export_tables(
            [small_df, large_df],
            filename=filename,
            max_rows_per_sheet=1000
        )
        
        filepath = tmp_path / filename
        wb = openpyxl.load_workbook(filepath)
        
        # Should have: Table1 (small), Table2_part1 (1000), Table2_part2 (500)
        assert len(wb.sheetnames) == 3
        assert "Table1" in wb.sheetnames
        assert "Table2_part1" in wb.sheetnames
        assert "Table2_part2" in wb.sheetnames
    
    def test_export_respects_default_max_rows(self, exporter, tmp_path):
        """Test that default max_rows_per_sheet is 1,000,000"""
        # Create DataFrame with 999,999 rows (should fit in one sheet)
        df = pd.DataFrame({'id': range(999999)})
        
        filename = "default_max.xlsx"
        exporter.export_tables(df, filename=filename)
        
        filepath = tmp_path / filename
        wb = openpyxl.load_workbook(filepath)
        
        # Should create only 1 sheet
        assert len(wb.sheetnames) == 1


class TestExportTablesVerboseMode(TestFileExporter):
    """Tests for verbose output during export_tables"""
    
    def test_verbose_output(self, tmp_path, sample_df, capsys):
        """Test that verbose mode produces output"""
        exporter = FileExporter(output_dir=str(tmp_path), verbose=True)
        
        exporter.export_tables(sample_df, filename="test.xlsx")
        
        captured = capsys.readouterr()
        assert "[INFO]" in captured.out
        assert "Exporting" in captured.out
    
    def test_verbose_multiple_tables(self, tmp_path, multiple_dfs, capsys):
        """Test verbose output with multiple tables"""
        exporter = FileExporter(output_dir=str(tmp_path), verbose=True)
        
        exporter.export_tables(multiple_dfs, filename="test.xlsx")
        
        captured = capsys.readouterr()
        assert "3 table(s)" in captured.out
        assert "Table1" in captured.out
        assert "Table2" in captured.out
        assert "Table3" in captured.out


class TestExportTablesEdgeCases(TestFileExporter):
    """Tests for edge cases in export_tables"""
    
    def test_export_dataframe_with_1_million_rows(self, exporter, tmp_path):
        """Test exporting DataFrame at the exact max_rows limit"""
        df = pd.DataFrame({'id': range(1000000)})
        
        filename = "exactly_1m.xlsx"
        exporter.export_tables(df, filename=filename, max_rows_per_sheet=1000000)
        
        filepath = tmp_path / filename
        wb = openpyxl.load_workbook(filepath)
        
        # Should fit in exactly 1 sheet
        assert len(wb.sheetnames) == 1
    
    def test_export_dataframe_with_1_million_plus_1_rows(self, exporter, tmp_path):
        """Test that 1,000,001 rows creates 2 sheets"""
        df = pd.DataFrame({'id': range(1000001)})
        
        filename = "over_1m.xlsx"
        exporter.export_tables(df, filename=filename, max_rows_per_sheet=1000000)
        
        filepath = tmp_path / filename
        wb = openpyxl.load_workbook(filepath)
        
        # Should create 2 sheets
        assert len(wb.sheetnames) == 2
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames
    
    def test_export_empty_dataframe(self, exporter, tmp_path):
        """Test exporting empty DataFrame"""
        df = pd.DataFrame()
        
        filename = "empty.xlsx"
        exporter.export_tables(df, filename=filename)
        
        filepath = tmp_path / filename
        assert filepath.exists()
    
    def test_export_with_special_characters_in_data(self, exporter, tmp_path):
        """Test exporting DataFrame with special characters"""
        df = pd.DataFrame({
            'name': ['José', 'Müller', 'François'],
            'city': ['São Paulo', 'München', 'Montréal']
        })
        
        filename = "special_chars.xlsx"
        exporter.export_tables(df, filename=filename)
        
        filepath = tmp_path / filename
        df_read = pd.read_excel(filepath)
        
        pd.testing.assert_frame_equal(df_read, df)
    
    def test_sheet_name_truncation(self, exporter, tmp_path):
        """Test that very long sheet names are truncated to 31 chars"""
        # Create many tables to force high numbers in sheet names
        dfs = [pd.DataFrame({'col': [i]}) for i in range(12)]
        
        filename = "many_tables.xlsx"
        exporter.export_tables(dfs, filename=filename)
        
        filepath = tmp_path / filename
        wb = openpyxl.load_workbook(filepath)
        
        # Verify all sheet names are <= 31 characters (Excel limit)
        for sheet_name in wb.sheetnames:
            assert len(sheet_name) <= 31


class TestExportTablesIntegration(TestFileExporter):
    """Integration tests combining read_all() and export_tables()"""
    
    def test_workflow_csv_to_excel(self, tmp_path):
        """Test reading CSV and exporting with export_tables"""
        from pandas_toolkit.io import CSVReader
        
        # Create CSV file
        csv_file = tmp_path / "input.csv"
        df_original = pd.DataFrame({
            'id': [1, 2, 3],
            'name': ['A', 'B', 'C']
        })
        df_original.to_csv(csv_file, index=False)
        
        # Read with read_all()
        reader = CSVReader()
        tables = reader.read_all(str(csv_file))
        
        # Should be a list with 1 DataFrame
        assert isinstance(tables, list)
        assert len(tables) == 1
        
        # Export with export_tables()
        exporter = FileExporter(output_dir=str(tmp_path))
        exporter.export_tables(tables, filename="output.xlsx")
        
        # Verify
        output_file = tmp_path / "output.xlsx"
        assert output_file.exists()
        df_read = pd.read_excel(output_file)
        pd.testing.assert_frame_equal(df_read, df_original)
